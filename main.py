import os
import platform
import threading
import subprocess

import pandas as pd
import tkinter as tk

from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.worksheet import Worksheet


# Initial Gathering
rootFolder = os.path.abspath(".")
outputFile = f"{rootFolder}/outputs/[{datetime.now().strftime('%d%m%Y%-H%M%S')}]Consolidated Time Data.xlsx"
templateFile = f"{rootFolder}/Files/Template.xlsx"

thinBorder = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

templateWB = load_workbook(templateFile)
tabs = templateWB.sheetnames

# Function to process the file (simulating heavy work)
def processFile(inputFile, additionalInfo, outputFileLabel, completeButton, sheetName, existingFile):
    try:
        inputData = pd.read_excel(inputFile, sheet_name=None)


        if existingFile == "":
            WS = templateWB[sheetName]
            useTemplate(WS, additionalInfo, inputData)
        else:
            existingWB = load_workbook(existingFile)
            WS = existingWB[sheetName]
            useExisting(WS, additionalInfo, inputData)

        # Update GUI to show completion
        showCompletionPopup(outputFile)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
    finally:
        submitButton.config(state=tk.NORMAL, text="Submit")

# Template Updater
def useTemplate(templateWS: Worksheet, additionalInfo: list, inputData: pd.DataFrame):
    additionalFields = [
        [2, 4],
        [3, 4],
        [6, 4],
        [7, 4],
        [8, 4]
    ]

    for index, info in enumerate(additionalInfo):
        templateWS.cell(additionalFields[index][0], additionalFields[index][1], info)
    
    dt = templateWS.cell(row=13, column=5).value
    totalCubes = templateWS.cell(row=13, column=6).value
    potHours = templateWS.cell(row=13, column=7).value
    potCubes = templateWS.cell(row=13, column=8).value
    lostHours = templateWS.cell(row=13, column=9).value
    lostProd = templateWS.cell(row=13, column=10).value

    currentRow = 13
    for index, sheet in enumerate(inputData):
        sheetData = inputData[sheet]

        if sheetData.values[3][2].lower() == "start" and sheetData.values[3][3].lower() == "stop":
            timeData = []
            dates = []
            for row in sheetData.values[4:]:
                if row[2] != row[2]:
                    continue
                try:
                    timeData.append(float(row[2]))
                    timeData.append(float(row[3]))
                    dates.append(row[0])
                except ValueError:
                    continue

            templateWS.insert_rows(currentRow)
        
            templateWS.cell(row=currentRow, column=2, value=sheet)
            templateWS.cell(row=currentRow, column=2).border = thinBorder
            templateWS.cell(row=currentRow, column=3, value=min(timeData))
            templateWS.cell(row=currentRow, column=3).border = thinBorder
            templateWS.cell(row=currentRow, column=4, value=max(timeData))
            templateWS.cell(row=currentRow, column=4).border = thinBorder
            templateWS.cell(row=currentRow, column=5, value=dt.replace("13", f"{currentRow}"))
            templateWS.cell(row=currentRow, column=5).border = thinBorder
            templateWS.cell(row=currentRow, column=6, value=totalCubes.replace("13", f"{currentRow}"))
            templateWS.cell(row=currentRow, column=6).border = thinBorder
            templateWS.cell(row=currentRow, column=7, value=potHours.replace("13", f"{currentRow}"))
            templateWS.cell(row=currentRow, column=7).border = thinBorder
            templateWS.cell(row=currentRow, column=8, value=potCubes.replace("13", f"{currentRow}"))
            templateWS.cell(row=currentRow, column=8).border = thinBorder
            templateWS.cell(row=currentRow, column=9, value=lostHours.replace("13", f"{currentRow}"))
            templateWS.cell(row=currentRow, column=9).border = thinBorder
            templateWS.cell(row=currentRow, column=10, value=lostProd.replace("13", f"{currentRow}"))
            templateWS.cell(row=currentRow, column=10).border = thinBorder

            currentRow += 1
    
    print(currentRow)

    templateWS.delete_rows(currentRow)

    dtSum = templateWS.cell(row=currentRow, column=5).value
    cubesSum = templateWS.cell(row=currentRow, column=6).value
    potHoursSum = templateWS.cell(row=currentRow, column=7).value
    potCubesSum = templateWS.cell(row=currentRow, column=8).value
    lostHoursSum = templateWS.cell(row=currentRow, column=9).value
    lostCubesSum = templateWS.cell(row=currentRow, column=10).value

    print(dtSum)
    # print(dtCheck)

    templateWS.cell(row=currentRow, column=5, value=dtSum.replace("13)", f"{currentRow-1})"))
    templateWS.cell(row=currentRow, column=5).border = thinBorder
    templateWS.cell(row=currentRow, column=6, value=cubesSum.replace("13)", f"{currentRow-1})"))
    templateWS.cell(row=currentRow, column=6).border = thinBorder
    templateWS.cell(row=currentRow, column=7, value=potHoursSum.replace("13)", f"{currentRow-1})"))
    templateWS.cell(row=currentRow, column=7).border = thinBorder
    templateWS.cell(row=currentRow, column=8, value=potCubesSum.replace("13)", f"{currentRow-1})"))
    templateWS.cell(row=currentRow, column=8).border = thinBorder
    templateWS.cell(row=currentRow, column=9, value=lostHoursSum.replace("13)", f"{currentRow-1})"))
    templateWS.cell(row=currentRow, column=9).border = thinBorder
    templateWS.cell(row=currentRow, column=10, value=lostCubesSum.replace("13)", f"{currentRow-1})"))
    templateWS.cell(row=currentRow, column=10).border = thinBorder

    templateWB.save(outputFile)

def useExisting(ws: Worksheet, additionalInfo: list, inputData: pd.DataFrame):
    additionalFields = [
        [2, 4],
        [3, 4],
        [6, 4],
        [7, 4],
        [8, 4]
    ]

    for index, info in enumerate(additionalInfo):
        ws.cell(additionalFields[index][0], additionalFields[index][1], info)
    
    dt = ws.cell(row=13, column=5).value
    totalCubes = ws.cell(row=13, column=6).value
    potHours = ws.cell(row=13, column=7).value
    potCubes = ws.cell(row=13, column=8).value
    lostHours = ws.cell(row=13, column=9).value
    lostProd = ws.cell(row=13, column=10).value

    currentRow = 13
    for index, sheet in enumerate(inputData):
        sheetData = inputData[sheet]

        if sheetData.values[3][2].lower() == "start" and sheetData.values[3][3].lower() == "stop":
            timeData = []
            dates = []
            for row in sheetData.values[4:]:
                if row[2] != row[2]:
                    continue
                try:
                    timeData.append(float(row[2]))
                    timeData.append(float(row[3]))
                    dates.append(row[0])
                except ValueError:
                    continue
            
            ws.cell(row=currentRow, column=2, value=sheet)
            ws.cell(row=currentRow, column=2).border = thinBorder
            ws.cell(row=currentRow, column=3, value=min(timeData))
            ws.cell(row=currentRow, column=3).border = thinBorder
            ws.cell(row=currentRow, column=4, value=max(timeData))
            ws.cell(row=currentRow, column=4).border = thinBorder
            ws.cell(row=currentRow, column=5, value=dt.replace("13", f"{currentRow}"))
            ws.cell(row=currentRow, column=5).border = thinBorder
            ws.cell(row=currentRow, column=6, value=totalCubes.replace("13", f"{currentRow}"))
            ws.cell(row=currentRow, column=6).border = thinBorder
            ws.cell(row=currentRow, column=7, value=potHours.replace("13", f"{currentRow}"))
            ws.cell(row=currentRow, column=7).border = thinBorder
            ws.cell(row=currentRow, column=8, value=potCubes.replace("13", f"{currentRow}"))
            ws.cell(row=currentRow, column=8).border = thinBorder
            ws.cell(row=currentRow, column=9, value=lostHours.replace("13", f"{currentRow}"))
            ws.cell(row=currentRow, column=9).border = thinBorder
            ws.cell(row=currentRow, column=10, value=lostProd.replace("13", f"{currentRow}"))
            ws.cell(row=currentRow, column=10).border = thinBorder

            currentRow += 1
            
# Function to display a pop-up when processing is complete
def showCompletionPopup(outputFile):
    popup = tk.Toplevel(root)
    popup.title("Processing Complete")
    popup.geometry("500x250")
    
    # Completion message
    tk.Label(popup, text="Processing Complete!", font=("Helvetica", 14)).pack(pady=10)
    tk.Label(popup, text=f"Output File: {outputFile}", wraplength=250).pack(pady=5)
    
    # Open file button
    tk.Button(popup, text="Open File", command=lambda: openFile(outputFile)).pack(pady=5)
    
    # Close button for the popup
    tk.Button(popup, text="Close", command=popup.destroy).pack(pady=5)

# Open the file with the default application
def openFile(filePath):
    if os.path.exists(filePath):
        systemName = platform.system()
        try:
            if systemName == "Darwin":  # macOS
                subprocess.call(["open", filePath])
            elif systemName == "Windows":  # Windows
                os.startfile(filePath)
            else:  # Linux and others
                subprocess.call(["xdg-open", filePath])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open the file: {e}")
    else:
        messagebox.showerror("Error", "File does not exist!")


# Handle the submit action
def onSubmit():
    inputFile = filePathEntry.get()
    additionalInfo = [entry.get() for entry in additionalInfoEntries]
    dropdownValue = dropdownVar.get()
    useExisting = existingFileVar.get()
    existingFile = existingFileEntry.get()

    if not inputFile:
        messagebox.showerror("Error", "Please select an Excel file.")
        return
    
    if any(not info for info in additionalInfo) and not useExisting:
        messagebox.showerror("Error", "Please fill out all additional information fields.")
        return
    
    if not dropdownValue:
        messagebox.showerror("Error", "Please select a value from the dropdown menu.")
        return

    if useExisting and not existingFile:
        messagebox.showerror("Error", "Please select an existing file.")
        return

    # Show busy state
    submitButton.config(state=tk.DISABLED, text="Processing...")
    # Process the file in a separate thread to avoid freezing the GUI
    thread = threading.Thread(target=processFile, args=(inputFile, additionalInfo, outputFileLabel, completeButton, dropdownValue, existingFile))
    thread.start()

# Browse for a file
def browseFile():
    filePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
    if filePath:
        filePathEntry.delete(0, tk.END)
        filePathEntry.insert(0, filePath.split("/")[-1])

# Browse for an existing file
def browseExistingFile():
    filePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
    if filePath:
        existingFileEntry.delete(0, tk.END)
        existingFileEntry.insert(0, filePath)

def toggleExistingFileInput():
    if existingFileVar.get():
        existingFileLabel.grid()
        existingFileEntry.grid()
        existingFileBrowseButton.grid()
    else:
        existingFileLabel.grid_remove()
        existingFileEntry.grid_remove()
        existingFileBrowseButton.grid_remove()

def closeApp():
    root.destroy()

# Create the GUI
root = tk.Tk()
root.title("ADT Time Tracker and Production Calculator")

# Checkbox for using an existing file
existingFileVar = tk.BooleanVar()
useExistingCheckbox = tk.Checkbutton(root, text="Use Existing File", variable=existingFileVar, command=toggleExistingFileInput)
useExistingCheckbox.grid(row=2, column=3, columnspan=3, pady=10)

# Existing file input (initially hidden)
existingFileLabel = tk.Label(root, text="Select Existing File:")
existingFileLabel.grid(row=3, column=3, padx=10, pady=10, sticky="e")
existingFileLabel.grid_remove()

existingFileEntry = tk.Entry(root, width=40)
existingFileEntry.grid(row=3, column=4, padx=10, pady=10)
existingFileEntry.grid_remove()

existingFileBrowseButton = tk.Button(root, text="Browse", command=browseExistingFile)
existingFileBrowseButton.grid(row=3, column=5, padx=10, pady=10)
existingFileBrowseButton.grid_remove()

# File selection
filePathLabel = tk.Label(root, text="Select Excel File:")
filePathLabel.grid(row=0, column=0, padx=10, pady=10, sticky="e")

filePathEntry = tk.Entry(root, width=40)
filePathEntry.grid(row=0, column=1, padx=10, pady=10)

browseButton = tk.Button(root, text="Browse", command=browseFile)
browseButton.grid(row=0, column=2, padx=10, pady=10)

additionalInfoEntries = []
additionalInfoLabels = [
    "Number of ADT's per Team:", 
    "Rate per ADT Team:", 
    "Shifts to Date:",
    "Days to Date:",
    "Production Hours per Day"
]
for i in range(5):
    label = tk.Label(root, text=additionalInfoLabels[i])
    label.grid(row=i + 1, column=0, padx=10, pady=10, sticky="e")

    entry = tk.Entry(root, width=40)
    entry.grid(row=i + 1, column=1, padx=10, pady=10, columnspan=1)
    additionalInfoEntries.append(entry)

# Dropdown menu
dropdownLabel = tk.Label(root, text="Select Month:")
dropdownLabel.grid(row=1, column=3, padx=10, pady=10, sticky="e")

dropdownVar = tk.StringVar()
dropdownMenu = ttk.Combobox(root, textvariable=dropdownVar, state="readonly")
dropdownMenu["values"] = tabs
dropdownMenu.grid(row=1, column=4, padx=10, pady=10)
dropdownMenu.current(0)  # Set the default value to the first option

# Submit button
submitButton = tk.Button(root, text="Submit", command=onSubmit)
submitButton.grid(row=6, column=1, pady=10)

# Close button
closeButton = tk.Button(root, text="Close", command=closeApp)
closeButton.grid(row=6, column=2, pady=10)

# Output status
outputFileLabel = tk.Label(root, text="")
outputFileLabel.grid(row=7, column=0, columnspan=3, pady=10)

# Open completed file button (initially hidden)
completeButton = tk.Button(root, text="Open File", state=tk.NORMAL)
completeButton.grid(row=8, column=1, pady=10)
completeButton.grid_remove()

# Run the GUI
root.mainloop()